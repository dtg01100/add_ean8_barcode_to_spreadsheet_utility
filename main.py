#!/usr/bin/env python3

import shutil
import warnings
import barcode
import openpyxl
import openpyxl.utils
from openpyxl.drawing.image import Image as OpenPyXlImage
from PIL import Image as pil_Image
import ImageOps as pil_ImageOps
import math
from barcode.writer import ImageWriter
import tkinter
import tkinter.ttk
import tkinter.filedialog
import tkinter.messagebox
import tempfile
import argparse
import textwrap
import threading
import platform
import os
import configparser
import appdirs
import re
import io
import barcode.pybarcode
import logging
from contextlib import redirect_stdout

version = '1.7.1'

appname = "Barcode Insert Utility"

supported_barcode_types = ['code39', 'ean8', 'ean13', 'UPC']

config_folder = appdirs.user_data_dir(appname)
try:
    os.makedirs(config_folder)
except FileExistsError:
    pass
settings_file_path = os.path.join(config_folder, 'barcode insert utility settings.cfg')

# read launch options
launch_options = argparse.ArgumentParser()
launch_options.add_argument('-d', '--debug', action='store_true', help="print debug output to stdout")
launch_options.add_argument('-l', '--log', action='store_true', help="write stdout to log file")
launch_options.add_argument('--keep_barcodes_in_cwd', action='store_true',
                            help="temp folder in working directory")
launch_options.add_argument('--keep_barcode_files', action='store_true', help="don't delete temp files")
launch_options.add_argument('--reset_configuration', action='store_true', help="remove configuration file")
args = launch_options.parse_args()

if args.reset_configuration:  # remove configuration file if reset_configuration flag is set
    try:
        os.remove(settings_file_path)
    except FileNotFoundError:
        pass

config = configparser.RawConfigParser()

if not os.path.exists(settings_file_path):
    config.add_section('settings')
    config.set('settings', 'initial_input_folder', os.path.expanduser('~'))
    config.set('settings', 'initial_output_folder', os.path.expanduser('~'))
    config.set('settings', 'barcode_dpi', '120')
    config.set('settings', 'barcode_module_height', '5')
    config.set('settings', 'barcode_border', '0')
    config.set('settings', 'barcode_font_size', '6')
    config.set('settings', 'input_data_column', 'B')
    config.set('settings', 'barcode_output_column', 'A')
    config.set('settings', 'barcode type', 'code39')
    config.set('settings', 'pad ean barcodes', "False")
    with open(settings_file_path, 'w', encoding='utf8') as configfile:
        config.write(configfile)

config.read(settings_file_path)  # open config file

root_window = tkinter.Tk()

root_window.title("Barcode Insert Utility " + version)

# this builds a list of the launch flags
flags_list_string = "Flags="
flags_count = 0
if args.debug:
    flags_list_string += "(Debug)"
    flags_count += 1
if args.log:
    flags_list_string += "(Logged)"
    flags_count += 1
if args.keep_barcodes_in_cwd:
    flags_list_string += "(Barcodes In Working Directory)"
    flags_count += 1
if args.keep_barcode_files:
    flags_list_string += "(Keep Barcodes)"
    flags_count += 1
if args.reset_configuration:
    flags_list_string += "(Reset Configuration)"
    flags_count += 1

old_workbook_path = ""
new_workbook_path = ""

program_launch_cwd = os.getcwd()

process_workbook_keep_alive = True


column_letter_list = []
column_count = 0
while column_count < 200:
    column_count += 1
    column_letter = openpyxl.utils.get_column_letter(column_count)
    column_letter_list.append(column_letter)
column_letter_tuple = tuple(column_letter_list)

barcode_type_variable = tkinter.StringVar()
pad_ean_option = tkinter.BooleanVar()


def invalid_configuration_error():
    root_window.withdraw()
    tkinter.messagebox.showerror(title="Batch File Sender Version " + version,
                                 message="Configuration file is broken, "
                                         "relaunch program with the option '--reset_configuration'")
    raise SystemExit


# set initial variables for configuration file test
barcode_dpi_test = None
barcode_module_height_test = None
barcode_border_test = None
barcode_font_size_test = None
barcode_type_test = None
invalid_configuration = False
input_barcode_test_column = None
output_barcode_test_column = None

# check to see if the following four are integers
try:
    barcode_dpi_test = config.getint('settings', 'barcode_dpi')
    barcode_module_height_test = config.getint('settings', 'barcode_module_height')
    barcode_border_test = config.getint('settings', 'barcode_border')
    barcode_font_size_test = config.getint('settings', 'barcode_font_size')
except ValueError:
    invalid_configuration_error()

try:
    input_barcode_test_column = config.get('settings', 'input_data_column')
    output_barcode_test_column = config.get('settings', 'barcode_output_column')
    barcode_type_test = config.get('settings', 'barcode type')
    _ = config.getboolean('settings', 'pad ean barcodes')
except (configparser.NoOptionError, ValueError):
    invalid_configuration_error()

# check that values are in acceptable ranges
if input_barcode_test_column not in column_letter_list or output_barcode_test_column not in column_letter_list:
    invalid_configuration = True
if barcode_dpi_test not in range(120, 400):
    invalid_configuration = True
if barcode_module_height_test not in range(5, 50):
    invalid_configuration = True
if barcode_border_test not in range(0, 25):
    invalid_configuration = True
if barcode_font_size_test not in range(0, 15):
    invalid_configuration = True
if barcode_type_test not in supported_barcode_types:
    invalid_configuration = True
if invalid_configuration:  # if any of the previous values are incorrect, show an error dialog and close out
    invalid_configuration_error()

# the following sets internal open file limits
try:
    if platform.system() == 'Windows':
        import win32file

        file_limit = win32file._getmaxstdio()
    else:
        import resource

        soft, hard = resource.getrlimit(resource.RLIMIT_NOFILE)
        file_limit = soft
except Exception as error:
    warnings.warn("Getting open file limit failed with: " + str(error) + " setting internal file limit to 100")
    file_limit = 100

if args.log:
    class Logger():
        def __init__(self):
            self.logger = logging.getLogger('Barcode Insert Utility')
            self.logger.setLevel(logging.DEBUG)
            fh = logging.FileHandler('logfile.log')
            ch = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(message)s')
            fh.setFormatter(formatter)
            ch.setFormatter(formatter)
            self.logger.addHandler(fh)
            self.logger.addHandler(ch)

        def write(self, message):
            self.logger.debug(message)

        def flush(self):
            # this flush method is needed for python 3 compatibility.
            # this handles the flush command by doing nothing.
            # you might want to specify some extra behavior here.
            pass


    loghandler = Logger()

if args.debug:
    print(launch_options.parse_args())


# this is a wrapper function for print, so that we can have it only spam stdout when debug is set
def print_if_debug(string):
    if args.debug:
        loghandler.write(string)


print_if_debug("Barcode Insert Utility version " + version)

print_if_debug("File limit is: " + str(file_limit))

# get supported formats from pybarcode, in case debugging is required
with io.StringIO() as buf, redirect_stdout(buf):
    barcode.pybarcode.list_types(ImageWriter)
    output = buf.getvalue()
print_if_debug(output)


# this is the workbook selector, the code was a bit of an experiment and is a bit of a pain to debug.
# not enough of the logic is shared between the two codepaths to make it worth the complexity
# ill probably rewrite this at some point.
def select_folder_old_new_wrapper(selection):
    global old_workbook_path
    global new_workbook_path
    for child in size_spinbox_frame.winfo_children():
        child.configure(state=tkinter.DISABLED)
    new_workbook_selection_button.configure(state=tkinter.DISABLED)
    old_workbook_selection_button.configure(state=tkinter.DISABLED)
    if selection == "old":
        old_workbook_path_proposed = tkinter.filedialog.askopenfilename(
            initialdir=config.get('settings', 'initial_input_folder'),
            filetypes=[("Excel Spreadsheet", "*.xlsx")])
        file_is_xlsx = False
        try:
            if os.path.exists(old_workbook_path_proposed):
                config.set('settings', 'initial_input_folder', os.path.dirname(old_workbook_path_proposed))
                with open(settings_file_path, 'w', encoding='utf8') as configuration_file:
                    config.write(configuration_file)
                try:
                    openpyxl.load_workbook(old_workbook_path_proposed, read_only=True)
                    file_is_xlsx = True
                except Exception as file_test_open_error:
                    print(file_test_open_error)
        except TypeError:
            old_workbook_path_proposed = ''
        if os.path.exists(old_workbook_path_proposed) and file_is_xlsx is True:
            old_workbook_path = old_workbook_path_proposed
            old_workbook_path_wrapped = '\n'.join(textwrap.wrap(old_workbook_path, width=75, replace_whitespace=False))
            old_workbook_label.configure(text=old_workbook_path_wrapped, justify=tkinter.LEFT)
    else:
        new_workbook_path_proposed = tkinter.filedialog.asksaveasfilename(
            initialdir=config.get('settings', 'initial_output_folder'),
            initialfile=os.path.basename(old_workbook_path),
            defaultextension='.xlsx',
            filetypes=[("Excel Spreadsheet", "*.xlsx")])
        try:
            if os.path.exists(os.path.dirname(new_workbook_path_proposed)):
                new_workbook_path = new_workbook_path_proposed
                config.set('settings', 'initial_output_folder', os.path.dirname(new_workbook_path))
                with open(settings_file_path, 'w', encoding='utf8') as configuration_file:
                    config.write(configuration_file)
                new_workbook_path_wrapped = '\n'.join(
                    textwrap.wrap(new_workbook_path, width=75, replace_whitespace=False))
                new_workbook_label.configure(text=new_workbook_path_wrapped, justify=tkinter.LEFT,
                                             background=old_workbook_label._root().cget('background'))
        except AttributeError:
            new_workbook_path = ''
    if os.path.exists(old_workbook_path) and os.path.exists(os.path.dirname(new_workbook_path)):
        process_workbook_button.configure(state=tkinter.NORMAL, text="Process Workbook")
    for child in size_spinbox_frame.winfo_children():
        child.configure(state=tkinter.NORMAL)
    set_spinbutton_state_read_only()
    new_workbook_selection_button.configure(state=tkinter.NORMAL)
    old_workbook_selection_button.configure(state=tkinter.NORMAL)


def generate_barcode(input_string, tempdir):
    border_size = int(border_spinbox.get())
    ean = barcode.get(barcode_type_variable.get(), input_string, writer=ImageWriter())
    # select output image size via dpi. internally, pybarcode renders as svg, then renders that as a png file.
    # dpi is the conversion from svg image size in mm, to what the image writer thinks is inches.
    ean.default_writer_options['dpi'] = int(dpi_spinbox.get())
    # module height is the barcode bar height in mm
    ean.default_writer_options['module_height'] = float(height_spinbox.get())
    # text distance is the distance between the bottom of the barcode, and the top of the text in mm
    ean.default_writer_options['text_distance'] = 1
    # font size is the text size in pt
    ean.default_writer_options['font_size'] = int(font_size_spinbox.get())
    # quiet zone is the distance from the ends of the barcode to the ends of the image in mm
    ean.default_writer_options['quiet_zone'] = 2
    # save barcode image with generated filename
    print_if_debug("generating barcode image")
    with tempfile.NamedTemporaryFile(dir=tempdir, suffix='.png', delete=False) as initial_temp_file_path:
        filename = ean.save(initial_temp_file_path.name[0:-4])
        print_if_debug("success, barcode image path is: " + filename)
        print_if_debug("opening " + str(filename) + " to add border")
        barcode_image = pil_Image.open(str(filename))  # open image as pil object
        print_if_debug("success")
        print_if_debug("adding barcode and saving")
        img_save = pil_ImageOps.expand(barcode_image, border=border_size,
                                        fill='white')  # add border around image
        width, height = img_save.size  # get image size of barcode with border
        # write out image to file
        with tempfile.NamedTemporaryFile(dir=tempdir, suffix='.png', delete=False) as final_barcode_path:
            img_save.save(final_barcode_path.name)
            print_if_debug("success, final barcode path is: " + final_barcode_path.name)
    return final_barcode_path.name, width, height

def interpret_barcode_string(upc_barcode_string):
    if not upc_barcode_string == '':
        if barcode_type_variable.get() == "ean13" or barcode_type_variable.get() == "ean8":
            try:
                _ = int(upc_barcode_string)  # check that "upc_barcode_string" can be cast to int
            except ValueError as exc:
                raise ValueError("Input contents are not an integer") from exc
        # select barcode type, specify barcode, and select image writer to save as png
        if barcode_type_variable.get() == "ean8":
            if pad_ean_option.get() is True:
                if len(upc_barcode_string) < 6:
                    upc_barcode_string = upc_barcode_string.rjust(6, '0')
                if len(upc_barcode_string) <= 7:
                    upc_barcode_string = upc_barcode_string.ljust(7, '0')
                else:
                    raise ValueError("Input contents are more than 7 characters")
            else:
                if len(upc_barcode_string) != 7:
                    raise ValueError("Input contents are not 7 characters")
        elif barcode_type_variable.get() == "ean13":
            if pad_ean_option.get() is True:
                if len(upc_barcode_string) < 11:
                    upc_barcode_string = upc_barcode_string.rjust(11, '0')
                if len(upc_barcode_string) <= 12:
                    upc_barcode_string = upc_barcode_string.ljust(12, '0')
                else:
                    raise ValueError("Input contents are more than 12 characters")
            else:
                if len(upc_barcode_string) != 12:
                    raise ValueError("Input contents are not 12 characters")
        elif barcode_type_variable.get() == "UPC":
            if pad_ean_option.get() is True:
                if len(upc_barcode_string) < 10:
                    upc_barcode_string = upc_barcode_string.rjust(11, '0')
                if len(upc_barcode_string) <= 11:
                    upc_barcode_string = upc_barcode_string.ljust(12, '0')
                else:
                    raise ValueError("Input contents are more than 11 characters")
            else:
                if len(upc_barcode_string) != 11:
                    raise ValueError("Input contents are not 11 characters")
        elif barcode_type_variable.get() == "code39":
            upc_barcode_string = upc_barcode_string.upper()
            upc_barcode_string = re.sub('[^A-Z0-9./*$%+\- ]+', ' ', upc_barcode_string)
        return upc_barcode_string
    raise ValueError("Input is empty")

def do_process_workbook():
    # this is called as a background thread to ensure the interface is responsive
    print_if_debug("creating temp directory")
    if not args.keep_barcodes_in_cwd:
        tempdir = tempfile.mkdtemp()
    else:
        temp_dir_in_cwd = os.path.join(program_launch_cwd, 'barcode images')
        os.mkdir(temp_dir_in_cwd)
        tempdir = temp_dir_in_cwd
    print_if_debug("temp directory created as: " + tempdir)
    progress_bar.configure(mode='indeterminate', maximum=100)
    progress_bar.start()
    progress_numbers.configure(text="opening workbook")
    wb = openpyxl.load_workbook(old_workbook_path)
    ws = wb.worksheets[0]
    progress_numbers.configure(text="testing workbook save")
    wb.save(new_workbook_path)
    count = 0
    save_counter = 0
    progress_bar.configure(maximum=ws.max_row, value=count)
    progress_numbers.configure(text=str(count) + "/" + str(ws.max_row))

    for _ in ws.iter_rows():  # iterate over all rows in current worksheet
        if not process_workbook_keep_alive:
            break
        try:
            count += 1
            progress_bar.configure(maximum=ws.max_row, value=count, mode='determinate')
            progress_numbers.configure(text=str(count) + "/" + str(ws.max_row))
            progress_bar.configure(value=count)
            # get code from column selected in input_colum_spinbox, on current row,
            # add a zeroes to the end if option is selected to make seven or 12 digits
            print_if_debug("getting cell contents on line number " + str(count))
            upc_barcode_string = str(ws[input_column_spinbox.get() + str(count)].value)
            print_if_debug("cell contents are: " + upc_barcode_string)
            upc_barcode_string = interpret_barcode_string(upc_barcode_string)
            generated_barcode_path, width, height = generate_barcode(upc_barcode_string, tempdir)
            # resize cell to size of image
            ws.column_dimensions[output_column_spinbox.get()].width = int(math.ceil(float(width) * .15))
            ws.row_dimensions[count].height = int(math.ceil(float(height) * .75))

            # open image with as openpyxl image object
            print_if_debug("opening " + generated_barcode_path + " to insert into output spreadsheet")
            img = OpenPyXlImage(generated_barcode_path)
            print_if_debug("success")
            # attach image to cell
            print_if_debug("adding image to cell")
            # add image to cell
            ws.add_image(img, anchor=output_column_spinbox.get() + str(count))
            save_counter += 1
            print_if_debug("success")
        except Exception as barcode_error:
            print_if_debug(barcode_error)
        # This save in the loop frees references to the barcode images,
        #  so that python's garbage collector can clear them
        if save_counter >= file_limit - 50:
            print_if_debug("saving intermediate workbook to free file handles")
            progress_bar.configure(mode='indeterminate', maximum=100)
            progress_bar.start()
            progress_numbers.configure(text=str(count) + "/" + str(ws.max_row) + " saving")
            wb.save(new_workbook_path)
            print_if_debug("success")
            save_counter = 1
            progress_numbers.configure(text=str(count) + "/" + str(ws.max_row))
    progress_bar.configure(value=0)
    print_if_debug("saving workbook to file")
    progress_bar.configure(mode='indeterminate', maximum=100)
    progress_bar.start()
    progress_numbers.configure(text="saving")
    wb.save(new_workbook_path)
    print_if_debug("success")
    if not args.keep_barcode_files:
        print_if_debug("removing temp folder " + tempdir)
        shutil.rmtree(tempdir)
        print_if_debug("success")


def process_workbook_thread():
    # this function handles setup and teardown of the process workbook thread
    global new_workbook_path
    process_errors = False
    global process_workbook_keep_alive
    process_workbook_keep_alive = True
    try:
        do_process_workbook()
    except (IOError, OSError):
        print("Error saving file")
        process_errors = True
        new_workbook_label.configure(text="Error saving, select another output file.", background='red')
    finally:
        progress_bar.stop()
        progress_bar.configure(maximum=100, value=0, mode='determinate')
        progress_numbers.configure(text="")
    new_workbook_path = ""
    if not process_errors:
        new_workbook_label.configure(text="No File Selected")


def process_workbook_command_wrapper():
    def kill_process_workbook():
        # this function sets the keep alive flag for the processing workbook thread to false,
        # this lets the thread exit gracefully, and clean up after itself
        global process_workbook_keep_alive
        process_workbook_keep_alive = False
        cancel_process_workbook_button.configure(text="Cancelling", state=tkinter.DISABLED)

    # this sets interface elements to disabled, and then saves config options to file
    new_workbook_selection_button.configure(state=tkinter.DISABLED)
    old_workbook_selection_button.configure(state=tkinter.DISABLED)
    config.set('settings', 'barcode_dpi', dpi_spinbox.get())
    config.set('settings', 'barcode_module_height', height_spinbox.get())
    config.set('settings', 'barcode_border', border_spinbox.get())
    config.set('settings', 'barcode_font_size', font_size_spinbox.get())
    config.set('settings', 'input_data_column', input_column_spinbox.get())
    config.set('settings', 'barcode_output_column', output_column_spinbox.get())
    config.set('settings', 'barcode type', barcode_type_variable.get())
    config.set('settings', 'pad ean barcodes', pad_ean_option.get())
    with open(settings_file_path, 'w', encoding='utf8') as configfile_before_processing:
        config.write(configfile_before_processing)
    for child in size_spinbox_frame.winfo_children():
        child.configure(state=tkinter.DISABLED)
    process_workbook_button.configure(state=tkinter.DISABLED, text="Processing Workbook")
    cancel_process_workbook_button = tkinter.ttk.Button(master=go_button_frame, command=kill_process_workbook,
                                                        text="Cancel")
    cancel_process_workbook_button.pack(side=tkinter.RIGHT)
    process_workbook_thread_object = threading.Thread(target=process_workbook_thread)
    process_workbook_thread_object.start()
    while process_workbook_thread_object.is_alive():
        root_window.update()
    cancel_process_workbook_button.destroy()
    new_workbook_selection_button.configure(state=tkinter.NORMAL)
    old_workbook_selection_button.configure(state=tkinter.NORMAL)
    for child in size_spinbox_frame.winfo_children():
        child.configure(state=tkinter.NORMAL)
    set_spinbutton_state_read_only()
    if process_workbook_keep_alive:
        process_workbook_button.configure(text="Done Processing Workbook")
    else:
        process_workbook_button.configure(text="Processing Workbook Canceled")


def generate_single_barcode():
    print("single barcode stub")

    if upc_entry.get() == '':
        return

    print_if_debug("creating temp directory")
    if not args.keep_barcodes_in_cwd:
        tempdir = tempfile.mkdtemp()
    else:
        temp_dir_in_cwd = os.path.join(program_launch_cwd, 'barcode images')
        os.mkdir(temp_dir_in_cwd)
        tempdir = temp_dir_in_cwd

    save_path = tkinter.filedialog.asksaveasfilename(
        initialdir=config.get('settings', 'initial_output_folder'),
        initialfile=upc_entry.get(),
        defaultextension='.png',
        filetypes=[("PNG Image File", "*.png")])

    try:
        if os.path.exists(os.path.dirname(save_path)):
            barcode_path, _, _ = generate_barcode(interpret_barcode_string(upc_entry.get()), tempdir)
            shutil.copyfile(barcode_path, save_path)
    except Exception as error:
        tkinter.messagebox.showerror(master=root_window, message=f"Failed to generate barcode image: {str(error)}")

    if not args.keep_barcode_files:
        print_if_debug("removing temp folder " + tempdir)
        shutil.rmtree(tempdir)
        print_if_debug("success")


both_workbook_frame = tkinter.ttk.Frame(root_window)
old_workbook_file_frame = tkinter.ttk.Frame(both_workbook_frame)
new_workbook_file_frame = tkinter.ttk.Frame(both_workbook_frame)
go_and_progress_frame = tkinter.ttk.Frame(root_window)
go_button_frame = tkinter.ttk.Frame(go_and_progress_frame)
progress_bar_frame = tkinter.ttk.Frame(go_and_progress_frame)
size_spinbox_frame = tkinter.ttk.Frame(root_window, relief=tkinter.GROOVE, borderwidth=2)


def set_spinbutton_state_read_only():
    # this sets all spinbuttons to readonly, this prevents invalid input from being inserted.
    # eventually, ill get a validator working and this will be unnecessary
    dpi_spinbox.configure(state='readonly')
    height_spinbox.configure(state='readonly')
    border_spinbox.configure(state='readonly')
    font_size_spinbox.configure(state='readonly')
    input_column_spinbox.configure(state='readonly')
    output_column_spinbox.configure(state='readonly')


barcode_type_menu = tkinter.ttk.OptionMenu(size_spinbox_frame, barcode_type_variable,
                                           config.get('settings', 'barcode type'), *supported_barcode_types)

pad_ean_checkbutton = tkinter.ttk.Checkbutton(size_spinbox_frame, text="Pad EAN Barcodes", variable=pad_ean_option,
                                              onvalue=True, offvalue=False)
pad_ean_option.set(config.getboolean('settings', 'pad ean barcodes'))

dpi_spinbox = tkinter.Spinbox(size_spinbox_frame, from_=120, to=400, width=3, justify=tkinter.RIGHT)
dpi_spinbox.delete(0, "end")
dpi_spinbox.insert(0, str(config.getint('settings', 'barcode_dpi')))

height_spinbox = tkinter.Spinbox(size_spinbox_frame, from_=5, to=50, width=3, justify=tkinter.RIGHT)
height_spinbox.delete(0, "end")
height_spinbox.insert(0, str(config.getint('settings', 'barcode_module_height')))

border_spinbox = tkinter.Spinbox(size_spinbox_frame, from_=0, to=25, width=3, justify=tkinter.RIGHT)
border_spinbox.delete(0, "end")
border_spinbox.insert(0, str(config.getint('settings', 'barcode_border')))

font_size_spinbox = tkinter.Spinbox(size_spinbox_frame, from_=0, to=15, width=3, justify=tkinter.RIGHT)
font_size_spinbox.delete(0, "end")
font_size_spinbox.insert(0, str(config.getint('settings', 'barcode_font_size')))

input_column_spinbox = tkinter.Spinbox(size_spinbox_frame, values=column_letter_tuple, width=3, justify=tkinter.RIGHT)
input_column_spinbox.delete(0, "end")
input_column_spinbox.insert(0, str(config.get('settings', 'input_data_column')))

output_column_spinbox = tkinter.Spinbox(size_spinbox_frame, values=column_letter_tuple, width=3, justify=tkinter.RIGHT)
output_column_spinbox.delete(0, "end")
output_column_spinbox.insert(0, str(config.get('settings', 'barcode_output_column')))

set_spinbutton_state_read_only()

old_workbook_selection_button = tkinter.ttk.Button(master=old_workbook_file_frame, text="Select Original Workbook",
                                                   command=lambda: select_folder_old_new_wrapper("old"))
old_workbook_selection_button.pack(anchor='w')

new_workbook_selection_button = tkinter.ttk.Button(master=new_workbook_file_frame, text="Select New Workbook",
                                                   command=lambda: select_folder_old_new_wrapper("new"))
new_workbook_selection_button.pack(anchor='w')

old_workbook_label = tkinter.ttk.Label(master=old_workbook_file_frame, text="No File Selected", relief=tkinter.SUNKEN)
new_workbook_label = tkinter.ttk.Label(master=new_workbook_file_frame, text="No File Selected", relief=tkinter.SUNKEN)

old_workbook_label.pack(anchor='w', padx=(1, 0))
new_workbook_label.pack(anchor='w', padx=(1, 0))

# create spinbox labels
barcode_type_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Barcode Type:", anchor=tkinter.E)
size_spinbox_dpi_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Barcode DPI:", anchor=tkinter.E)
size_spinbox_height_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Barcode Height:", anchor=tkinter.E)
border_spinbox_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Barcode Border:", anchor=tkinter.E)
font_size_spinbox_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Barcode Text Size:", anchor=tkinter.E)
input_column_spinbox_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Input Column:", anchor=tkinter.E)
output_column_spinbox_label = tkinter.ttk.Label(master=size_spinbox_frame, text="Output Column:", anchor=tkinter.E)

# insert spinbox labels into frame with grid packer
barcode_type_label.grid(row=0, column=0, sticky=tkinter.W + tkinter.E, pady=2)
size_spinbox_dpi_label.grid(row=1, column=0, sticky=tkinter.W + tkinter.E, pady=2)
size_spinbox_height_label.grid(row=2, column=0, sticky=tkinter.W + tkinter.E, pady=2)
border_spinbox_label.grid(row=3, column=0, sticky=tkinter.W + tkinter.E, pady=2)
font_size_spinbox_label.grid(row=1, column=2, sticky=tkinter.W + tkinter.E, pady=2)
input_column_spinbox_label.grid(row=2, column=2, sticky=tkinter.W + tkinter.E, pady=2)
output_column_spinbox_label.grid(row=3, column=2, sticky=tkinter.W + tkinter.E, pady=2)

# set labels as stretchable
barcode_type_label.columnconfigure(0, weight=1)
size_spinbox_dpi_label.columnconfigure(0, weight=1)
size_spinbox_height_label.columnconfigure(0, weight=1)
border_spinbox_label.columnconfigure(0, weight=1)
font_size_spinbox_label.columnconfigure(0, weight=1)
input_column_spinbox_label.columnconfigure(0, weight=1)
output_column_spinbox_label.columnconfigure(0, weight=1)

# insert spinboxes into frame with grid packer
barcode_type_menu.grid(row=0, column=1)
dpi_spinbox.grid(row=1, column=1, sticky=tkinter.E, pady=2, padx=(0, 2))
height_spinbox.grid(row=2, column=1, sticky=tkinter.E, pady=2, padx=(0, 2))
border_spinbox.grid(row=3, column=1, sticky=tkinter.E, pady=2, padx=(0, 2))
pad_ean_checkbutton.grid(row=0, column=2, columnspan=2, sticky=tkinter.E)
font_size_spinbox.grid(row=1, column=3, sticky=tkinter.E, pady=2)
input_column_spinbox.grid(row=2, column=3, sticky=tkinter.E, pady=2)
output_column_spinbox.grid(row=3, column=3, sticky=tkinter.E, pady=2)


sidebar_frame = tkinter.ttk.Frame(master=root_window)

single_barcode_frame = tkinter.Frame(master=sidebar_frame)
single_barcode_frame.grid(row=0, column=1)
show_sidebar = True
def toggle_single_process_sidebar():
    global show_sidebar
    if show_sidebar:
        single_barcode_frame.grid_remove()
        show_sidebar = False
    else:
        single_barcode_frame.grid()
        show_sidebar = True


toggle_sidebar_button = tkinter.ttk.Button(master=sidebar_frame, text=">", command=toggle_single_process_sidebar, width=1)
toggle_sidebar_button.grid(row=0, column=0, sticky=tkinter.N + tkinter.S)

toggle_single_process_sidebar()

upc_entry = tkinter.ttk.Entry(master=single_barcode_frame)
create_barcode = tkinter.ttk.Button(master=single_barcode_frame, text="Save Barcode...", command=generate_single_barcode)

upc_entry.pack()
create_barcode.pack()

process_workbook_button = tkinter.ttk.Button(master=go_button_frame, text="Select Workbooks",
                                             command=process_workbook_command_wrapper)

process_workbook_button.configure(state=tkinter.DISABLED)

process_workbook_button.pack(side=tkinter.LEFT)

progress_bar = tkinter.ttk.Progressbar(master=progress_bar_frame)
progress_bar.pack(side=tkinter.RIGHT)
progress_numbers = tkinter.ttk.Label(master=progress_bar_frame)
progress_numbers.pack(side=tkinter.LEFT)

if flags_count != 0:  # only show flags header when there is something to display
    tkinter.ttk.Label(root_window, text=flags_list_string).grid(row=0, column=0, columnspan=2)
old_workbook_file_frame.pack(anchor='w', pady=2)
new_workbook_file_frame.pack(anchor='w', pady=(3, 2))
both_workbook_frame.grid(row=1, column=0, sticky=tkinter.W, padx=5, pady=5)
both_workbook_frame.columnconfigure(0, weight=1)
size_spinbox_frame.grid(row=1, column=1, sticky=tkinter.E + tkinter.N, padx=5, pady=(8, 5))
size_spinbox_frame.columnconfigure(0, weight=1)
sidebar_frame.grid(row=1, column=2, sticky=tkinter.N + tkinter.S, padx=5, pady=(8, 5))
sidebar_frame.rowconfigure(0, weight=1)
go_button_frame.pack(side=tkinter.LEFT, anchor='w')
progress_bar_frame.pack(side=tkinter.RIGHT, anchor='e')
go_and_progress_frame.grid(row=2, column=0, columnspan=2, sticky=tkinter.W + tkinter.E, padx=5, pady=5)
go_and_progress_frame.columnconfigure(0, weight=1)
root_window.columnconfigure(0, weight=1)

root_window.minsize(400, root_window.winfo_height())
root_window.resizable(width=tkinter.FALSE, height=tkinter.FALSE)

root_window.mainloop()
